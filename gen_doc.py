import datetime
from openpyxl import Workbook

class Exam():
    def __init__(self, title='Title'):
        self.title = title
        self.new()

    def new(self):
        self.wb = Workbook()
        self.ws = self.wb.active

    def format_sheet(self, start=0, end=1):
        MARGIN_WIDTH = 6
        SPACE_WIDTH = 4
        FORMULAR_WIDTH = 15
        ROW_HEIGHT = 25

        self.ws['B1'] = self.title
        self.ws['B2'] = 'Time:'
        self.ws['F2'] = 'Score:'
        self.ws.merge_cells('B1:H1')

        # margin width
        self.ws.column_dimensions['A'].width = MARGIN_WIDTH
        self.ws.column_dimensions['I'].width = MARGIN_WIDTH

        self.ws.column_dimensions['C'].width = SPACE_WIDTH
        self.ws.column_dimensions['E'].width = SPACE_WIDTH
        self.ws.column_dimensions['G'].width = SPACE_WIDTH


        self.ws.column_dimensions['B'].width = FORMULAR_WIDTH
        self.ws.column_dimensions['D'].width = FORMULAR_WIDTH
        self.ws.column_dimensions['F'].width = FORMULAR_WIDTH
        self.ws.column_dimensions['H'].width = FORMULAR_WIDTH

        for col in range(start, end):
            print(col)
            self.ws.row_dimensions[col].height = ROW_HEIGHT

        for col in range(start, end):
            print(col, self.ws.row_dimensions[col].height)
            self.ws.row_dimensions[col].height = ROW_HEIGHT

    def add(self, i, dat):
        # row = (i + 3) / 4
        row = i / 4
        col = i % 4
        print('%d:%d, %s' % (row, col, dat))
        self.ws.cell(row+3, col*2+2).value = dat

    def save(self, file_name='sample.xlsx'):
        self.wb.save(file_name)
